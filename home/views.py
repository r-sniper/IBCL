from django.shortcuts import render

# Create your views here.
from openpyxl.styles import Alignment

from home.models import TShirtSize, Colour, Team, Player, Instruction
from django.db import IntegrityError
from django.http import HttpResponse
from openpyxl import Workbook


def home(request):
    t_shirt_size = TShirtSize.objects.filter(is_active=True)
    t_shirt_colour = Colour.objects.filter(is_active=True, is_used=False)

    instructions = Instruction.objects.all()[0];

    return render(request, 'base.html', {
        'player_size': range(1, 8),
        't_shirt_size': t_shirt_size,
        't_shirt_colour': t_shirt_colour,
        'instructions': instructions.instructions,
    })


def register(request):
    if request.method == "POST":
        team_name = request.POST.get('team_name')
        selected_colour_pk = int(request.POST.get("t_shirt_colour"))
        t_shirt_colour = Colour.objects.get(pk=selected_colour_pk)
        try:
            team_obj = Team.objects.create(team_name=team_name, t_shirt_colour=t_shirt_colour)
        except IntegrityError:
            colour_used = Colour.objects.filter(is_used=True)
            colours = [each.colour_name for each in colour_used]
            return HttpResponse(
                "Either Colour or Team name is already taken by other team<br> Following Colours are already used<br>" + str(
                    colours) + '<br> Press back button and correct it.<br> Thank you for cooperation')
        t_shirt_colour.is_used = True
        t_shirt_colour.save()
        team_obj.save()
        captain_number = int(request.POST.get("is_captain"))
        for i in range(1, 9):
            print(i)
            player_name = request.POST.get('player_name_' + str(i))
            print_name = request.POST.get('player_print_name_' + str(i))
            t_shirt_number = request.POST.get('player_t_shirt_number_' + str(i))
            t_shirt_size_pk = request.POST.get('player_t_shirt_size_' + str(i))
            t_shirt_obj = TShirtSize.objects.get(pk=t_shirt_size_pk)
            is_captain = False
            if captain_number == i:
                is_captain = True

            print(player_name)
            print(print_name)
            print(t_shirt_number)
            print(t_shirt_size_pk)
            print(t_shirt_obj)
            print(is_captain)
            if player_name != "" and print_name != "" and t_shirt_number != "" and player_name != None and print_name != None and t_shirt_number != None:
                player_obj = Player.objects.create(player_name=player_name, printing_name=print_name,
                                                   t_shirt_number=t_shirt_number, t_shirt_size=t_shirt_obj,
                                                   is_captain=is_captain, team=team_obj)

                player_obj.save()

        return render(request, 'success.html')


def generate_excel(request):
    wb = Workbook()
    ws1 = wb.active

    column_heading = ['Team Name', "Colour", 'Player 1', 'Player 2', 'Player 3', 'Player 4', 'Player 5', 'Player 6',
                      'Player 7', 'Player 8', 'Captain']

    column_heading = ['Team', 'Player', 'Captain', 'Jersey Name', 'Jersey Number', 'Jersey Size', 'Colour']

    ws1.append(column_heading)

    all_teams = Team.objects.all()
    row_counter = 2
    for each_team in all_teams:
        column_counter = 1
        all_player_team = each_team.player_set.all()
        ws1.cell(row=row_counter, column=column_counter).value = each_team.team_name
        ws1.cell(row=row_counter, column=column_counter + 6).value = each_team.t_shirt_colour.colour_name
        ws1.merge_cells(start_row=row_counter, start_column=column_counter, end_row=row_counter + len(all_player_team),
                        end_column=column_counter)
        ws1.merge_cells(start_row=row_counter, start_column=column_counter + 6,
                        end_row=row_counter + len(all_player_team),
                        end_column=column_counter + 6)

        ws1.cell(row=row_counter, column=column_counter).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row_counter, column=column_counter + 6).alignment = Alignment(horizontal="center",
                                                                                   vertical="center")

        for each_player in all_player_team:
            ws1.cell(row=row_counter, column=column_counter + 1).value = each_player.player_name
            ws1.cell(row=row_counter, column=column_counter + 3).value = each_player.printing_name
            ws1.cell(row=row_counter, column=column_counter + 4).value = each_player.t_shirt_number
            ws1.cell(row=row_counter, column=column_counter + 5).value = each_player.t_shirt_size.size

            if each_player.is_captain == True:
                ws1.cell(row=row_counter, column=column_counter + 2).value = "Y"
            row_counter += 1

        row_counter += 1

    wb.save(filename="abcd.xlsx")
    return HttpResponse("Generated Excel")


def slogan(request):
    if request.method == "GET":
        all_teams = Team.objects.filter(slogan_submitted=False)
        print(all_teams)
        return render(request, 'slogan_filling.html', {
            'all_teams': all_teams
        })
    else:
        team_id = request.POST.get('team_name')
        team_obj = Team.objects.get(pk=team_id)
        if team_obj.slogan_submitted == True:
            return HttpResponse("Team slogan Already submitted")
        else:
            team_obj.slogan = request.POST.get('slogan')
            team_obj.slogan_submitted = True
            team_obj.save()
            return HttpResponse(
                "Team Name:" + team_obj.team_name + "<br>" + "Slogan:" + team_obj.slogan + "<br> Successfully Registered")
